import zipfile
from openpyxl import load_workbook
from PyPDF2 import PdfReader
import shutil
import os

def test_archive_and_check_files():

    archive_name = "new_archive.zip"

    with zipfile.ZipFile(archive_name, 'w') as myzip:
        myzip.write("resources/file.xlsx")
        myzip.write("resources/file1.xls")
        myzip.write("resources/read.txt")
        myzip.write("resources/REST_API.pdf")

    if not os.path.isdir('tmp'):
        os.mkdir('tmp')
        shutil.move("new_archive.zip", os.path.join('tmp', "new_archive.zip"))

        with myzip.open("resources/REST_API.pdf") as pdf_file:
            pdf_file = PdfReader(pdf_file)
            page = pdf_file.pages[0]
            text = page.extract_text()
            assert "Краткая шпаргалка по запросам REST API" in text

        with myzip.open("resources/file.xlsx") as xlsx_file:
            xlsx_file = load_workbook(xlsx_file)
            sheet = xlsx_file.active
            value = sheet.cell(row=3, column=2).value
            assert value == "Получение токена"


        with myzip.open("resources/file1.xls") as xls_file:
            xls_file = load_workbook(xls_file)
            sheet = xls_file.active
            value = sheet.cell(row=3, column=2).value
            assert value == "Получение токена"

        with myzip.open("resources/read.txt") as txt_file:
            text = txt_file.read()
            assert b"pytest has never been associated with a security vulnerability" in text